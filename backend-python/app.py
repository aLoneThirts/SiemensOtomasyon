from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.responses import JSONResponse
import openpyxl
import os
import shutil
import uuid
import logging
import sqlite3
from typing import List, Dict, Optional
import uvicorn
from dotenv import load_dotenv

# .env dosyasini yukle
load_dotenv()

app = FastAPI(title="Siemens Excel API")

# ─── P0-4: API Key Authentication ─────────────────────────────────────────
# .env dosyasina API_SECRET_KEY=buraya-guclu-bir-key-yaz seklinde ekle
API_SECRET_KEY = os.environ.get("API_SECRET_KEY", "")

async def verify_api_key(x_api_key: str = Header(default="")):
    """Her istekte X-Api-Key header'ini kontrol eder."""
    if not API_SECRET_KEY:
        # ✅ FIX: API key tanımlı değilse sunucu başlatılmamalı
        raise HTTPException(status_code=500, detail="API_SECRET_KEY konfigüre edilmemiş! .env dosyasini kontrol edin.")
    if x_api_key != API_SECRET_KEY:
        raise HTTPException(status_code=401, detail="Yetkisiz erisim")


# ─── CORS Whitelist ────────────────────────────────────────────────────────
ALLOWED_ORIGINS = [
    "http://localhost:3000",
    "http://localhost:5173",
    # "https://nexledger.tufekci.com",  # Production domain eklenince ac
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["Content-Type", "Authorization", "X-Api-Key"],
)

# Nginx reverse proxy arkadasinda gercek IP tespiti icin
# Production'da allowed_hosts'a kendi domain'ini ekle
# app.add_middleware(
#     TrustedHostMiddleware,
#     allowed_hosts=["nexledger.tufekci.com", "localhost"]
# )

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ─── P0-5: SQLite Kalici Storage ──────────────────────────────────────────
DB_PATH = "urunler.db"

def init_db():
    """SQLite veritabanini olustur."""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS urunler (
            urun_kodu TEXT PRIMARY KEY,
            urun_adi TEXT DEFAULT '',
            fiyat REAL DEFAULT 0
        )
    """)
    conn.commit()
    conn.close()

# Uygulama baslarken DB'yi olustur
init_db()

def db_connect():
    """SQLite baglantisi dondurur."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ─── Endpoints ─────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {
        "success": True,
        "message": "Siemens Excel API Calisiyor",
        "python_version": "3.14+",
        "durum": "Aktif"
    }


@app.post("/excel/yukle")
async def excel_yukle(file: UploadFile = File(...), _=Depends(verify_api_key)):
    # P0-4a: Dosya tipi kontrolu
    allowed_extensions = (".xlsx", ".xls")
    original_name = file.filename or ""
    if not original_name.lower().endswith(allowed_extensions):
        raise HTTPException(
            status_code=400,
            detail="Sadece Excel dosyalari (.xlsx, .xls) kabul edilir."
        )

    try:
        # P0-4b: UUID dosya adi - path traversal onlemi
        safe_filename = f"{uuid.uuid4().hex}.xlsx"
        dosya_yolu = os.path.join(UPLOAD_DIR, safe_filename)
        with open(dosya_yolu, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Excel'i oku
        wb = openpyxl.load_workbook(dosya_yolu, data_only=True)
        sheet = wb.active

        # Basliklari bul
        basliklar = []
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            basliklar = [str(cell).lower() if cell else "" for cell in row]
            break

        # Sutun index'lerini bul
        urun_kodu_idx = -1
        urun_adi_idx = -1
        fiyat_idx = -1

        for i, baslik in enumerate(basliklar):
            if any(k in baslik for k in ["urun kodu", "product", "kod"]):
                urun_kodu_idx = i
            elif any(k in baslik for k in ["urun adi", "ad", "name"]):
                urun_adi_idx = i
            elif any(k in baslik for k in ["fiyat", "price", "tutar"]):
                fiyat_idx = i

        # Urunleri oku ve SQLite'a yaz
        conn = db_connect()
        eklenen = 0
        guncellenen = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if urun_kodu_idx >= 0 and row[urun_kodu_idx]:
                kod = str(row[urun_kodu_idx]).strip()
                adi = ""
                fiyat = 0.0

                if urun_adi_idx >= 0 and row[urun_adi_idx]:
                    adi = str(row[urun_adi_idx]).strip()

                if fiyat_idx >= 0 and row[fiyat_idx]:
                    try:
                        fiyat = float(row[fiyat_idx])
                    except (ValueError, TypeError):
                        pass

                # UPSERT: varsa guncelle, yoksa ekle
                cursor = conn.execute("SELECT 1 FROM urunler WHERE urun_kodu = ?", (kod,))
                if cursor.fetchone():
                    conn.execute(
                        "UPDATE urunler SET urun_adi = ?, fiyat = ? WHERE urun_kodu = ?",
                        (adi, fiyat, kod)
                    )
                    guncellenen += 1
                else:
                    conn.execute(
                        "INSERT INTO urunler (urun_kodu, urun_adi, fiyat) VALUES (?, ?, ?)",
                        (kod, adi, fiyat)
                    )
                    eklenen += 1

        conn.commit()
        conn.close()
        wb.close()

        toplam = eklenen + guncellenen
        return {
            "success": True,
            "toplam_urun": toplam,
            "eklenen": eklenen,
            "guncellenen": guncellenen,
            "mesaj": f"{toplam} urun islendi ({eklenen} yeni, {guncellenen} guncellendi)"
        }

    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Excel yukleme hatasi: {e}")
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "error": "Excel dosyasi islenirken bir hata olustu. Dosya formatini kontrol edin."
            }
        )


@app.get("/urun/ara/{aranan}")
async def urun_ara(aranan: str, _=Depends(verify_api_key)):
    if len(aranan) < 2:
        return {"sonuclar": []}

    try:
        conn = db_connect()
        cursor = conn.execute(
            "SELECT urun_kodu, urun_adi, fiyat FROM urunler WHERE LOWER(urun_kodu) LIKE ? LIMIT 20",
            (f"%{aranan.lower()}%",)
        )
        sonuclar = [dict(row) for row in cursor.fetchall()]
        conn.close()

        return {
            "success": True,
            "sonuclar": sonuclar,
            "toplam": len(sonuclar)
        }
    except Exception as e:
        logging.error(f"Urun arama hatasi: {e}")
        return {"success": False, "sonuclar": [], "toplam": 0}


@app.get("/urun/gruplar")
async def urun_gruplari(_=Depends(verify_api_key)):
    try:
        conn = db_connect()
        cursor = conn.execute("SELECT urun_kodu FROM urunler ORDER BY urun_kodu")
        gruplar: Dict[str, List[str]] = {}

        for row in cursor.fetchall():
            kod = row["urun_kodu"]
            if kod and len(kod) >= 3:
                baslangic = kod[:3].upper()
                if baslangic not in gruplar:
                    gruplar[baslangic] = []
                if len(gruplar[baslangic]) < 5:
                    gruplar[baslangic].append(kod)

        conn.close()

        return {
            "success": True,
            "toplam_grup": len(gruplar),
            "gruplar": gruplar
        }
    except Exception as e:
        logging.error(f"Urun gruplar hatasi: {e}")
        return {"success": False, "toplam_grup": 0, "gruplar": {}}


@app.get("/urun/sayisi")
async def urun_sayisi(_=Depends(verify_api_key)):
    """Toplam urun sayisini dondurur."""
    try:
        conn = db_connect()
        cursor = conn.execute("SELECT COUNT(*) as sayi FROM urunler")
        sayi = cursor.fetchone()["sayi"]
        conn.close()
        return {"success": True, "toplam": sayi}
    except Exception as e:
        logging.error(f"Urun sayisi hatasi: {e}")
        return {"success": False, "toplam": 0}


if __name__ == "__main__":
    print("Siemens Excel API Baslatiliyor - http://localhost:8000")
    uvicorn.run(app, host="0.0.0.0", port=8000)